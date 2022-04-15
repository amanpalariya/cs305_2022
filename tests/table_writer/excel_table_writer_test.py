from nbformat import write
import pytest

from main.table_writer.excel_table_writer import ExcelTableWriter


@pytest.fixture
def workbook_path():
    return "tests/resources/output.xlsx"

def test_write(workbook_path):
    header = ["Header 1", "Header 2"]
    rows = [
        ["Row 1 Col 0", "Row 1 Col 1"],
        ["Row 2 Col 0", "Row 2 Col 1"],
    ]
    writer = ExcelTableWriter(header, workbook_path, 'Sheet 1')
    writer.add_row(rows[0])
    assert writer.get_number_of_rows_added() == 1
    writer.add_row(rows[1])
    writer.close()

    import openpyxl
    print(workbook_path)
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb.active
    for i in range(2):
        assert sheet.cell(1, i+1).value == header[i]
    for i in range(2):
        for j in range(2):
            assert sheet.cell(i+2, j+1).value == rows[i][j]

def test_write_fail(workbook_path):
    header = ["Header 1", "Header 2"]
    writer = ExcelTableWriter(header, workbook_path, 'Sheet 1')
    try:
        writer.add_row(["Only 1 cell"])
        assert False
    except:
        assert True